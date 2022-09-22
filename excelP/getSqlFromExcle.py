import openpyxl


def getSqlFromExcel( filePaht ):

    workbook = openpyxl.load_workbook(filePaht)
    worksheet1 = workbook.get_sheet_by_name('Sheet1')
    rows = worksheet1.max_row

    for row in range(1,rows+1):
        for col in range(1,3):
            print( worksheet1.cell(row=row, column=col).value)








